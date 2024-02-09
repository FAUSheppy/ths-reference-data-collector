#!/usr/bin/python3

import datetime
import argparse
import datetime as dt
import dateutil.relativedelta
import os
import requests
import csv
import pyexcel.cookbook
import pyexcel
import openpyxl
import glob
import calendar

import locale
import fallback_csv
import platform

if platform.system() != "Linux":
    locale.setlocale(locale.LC_TIME, "de_DE.UTF-8")

CSV_DIR = "csvfiles"
CACHE_DIR = "cache"
CACHE_FILE_TEMPLATE  = "cache_{}_{}_{}.data"
NFF_URL_TIMEFORMAT   = "%d.%m.%Y"
NFF_INPUT_TIMEFORMAT = "%d.%m.%Y %H:%M"
TIME_FMT_DE_EXCEL    = "%d.%m.%Y %H:%M"

DATE_STYLE = openpyxl.styles.NamedStyle(name='custom_de_datetime', number_format='DD.MM.YYYY HH:MM')

OUTSIDE_DATA_URL     = "http://umweltdaten.nuernberg.de/csv/wetterdaten/messstation-nuernberg-flugfeld/archiv/csv-export/SUN/nuernberg-flugfeld/{dtype}/individuell/{fromDate}/{toDate}/export.csv"

headerMappings = {
            "time"                  : "Datum/Zeit",
            "lufttemperatur-aussen" : "Temperatur\n[°C]",
            "kelvin"                : "Temperatur\n[K]" ,
            "luftfeuchte"           : "rel. Luftfeuchte\n[%]",
            "luftdruck"             : "Luftdruck\n[mbar]",
            "windgeschwindigkeit"   : "Windgeschwindig-\nkeit\n[m/s]",
            "windrichtung"          : "Windrichtung\nN=0, O=90,\nS=180, W=270",
            "niederschlagsmenge"    : "Niederschlag\n[mm = L/m2]" }

dtypes  = [ "lufttemperatur-aussen", "luftfeuchte", "luftdruck", "windgeschwindigkeit", "windrichtung", "niederschlagsmenge" ]

def downloadFlugfeldData(fromTime, toTime, dtype):

    # prepare strings #
    cacheDir    = CACHE_DIR
    fromTimeStr = fromTime.strftime(NFF_URL_TIMEFORMAT)
    toTimeStr   = toTime.strftime(NFF_URL_TIMEFORMAT)
    cacheFile   = CACHE_FILE_TEMPLATE.format(dtype, fromTimeStr, toTimeStr)
    fullpath    = os.path.join(cacheDir, cacheFile)

    # remove empty placeholder cache file #
    if os.path.getsize(fullpath) == 0:
        os.remove(fullpath)

    # check for cache file
    content = None
    if not os.path.isfile(fullpath):
        url = OUTSIDE_DATA_URL.format(dtype=dtype, fromDate=fromTimeStr, toDate=toTimeStr)
        r = requests.get(url)
        content = r.content.decode('utf-8', "ignore") # ignore bad bytes

        # check response code #
        if r.status_code != 200 or "nicht gefunden" in r.text.lower():
            content = fallback_csv.generate("./dwd", fromTime, toTime, cacheFile, dtype)
        else:
            content = r.content.decode('utf-8', "ignore") # ignore bad bytes

        # cache data
        if not os.path.isdir(cacheDir):
            os.mkdir(cacheDir)
        with open(fullpath, 'w') as f:
            f.write(content)

    if os.path.isfile(fullpath):
        with open(fullpath) as f:
            content = f.read()

    return content

def checkLastMonths(backwardsMonths=6):


    today = dt.datetime.today()
    monthsToCheck = [ today.month - x for x in range(0, backwardsMonths)  ]
    monthsToCheckFixed = list(map(lambda x: x if x > 0 else x + 12, monthsToCheck))

    for monthNumber in monthsToCheckFixed:

        fullContentDict = dict()
        year = today.year
        if monthNumber > today.month:
            year = today.year - 1
        start = dt.datetime(year=year, month=monthNumber, day=1)
        end = start + dateutil.relativedelta.relativedelta(months=+1, seconds=-1)


        # check special cases #
        if end > today:
            end = today - dt.timedelta(days=4)
            if start > end:
                return ""

        for dtype in dtypes:
            content = downloadFlugfeldData(start, end, dtype)
            dataList = parse(content, dtype)
            for d in dataList:
                if d.time in fullContentDict:
                    fullContentDict[d.time] += [d]
                else:
                    fullContentDict.update({ d.time : [d] })

        # parse and dump
        mname = calendar.month_name[monthNumber]
        if monthNumber == 3:
            mname = "März" # fix german months

        csvOut = os.path.join(CSV_DIR, 'Wetterdaten-{}-{}.csv'.format(mname, year))
        with open(csvOut, 'w', newline='', encoding="utf-8") as file:

            fieldnames = list(headerMappings.values())
            writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
            writer.writeheader()

            for key in fullContentDict.keys():
                rowdict = { headerMappings["time"] : key }
                for data in fullContentDict[key]:
                    rowdict.update({ headerMappings[data.dtype] : data.value })

                    # calc kelvin if temp #
                    if data.dtype == "lufttemperatur-aussen":
                        rowdict.update({ headerMappings["kelvin"] : data.value + 273 })

                writer.writerow(rowdict)

def parse(content, dtype):
    skipBecauseFirstLine = True
    dataList = []
    for l in content.split("\n"):
        if not ";" in l:
            continue
        elif not l.strip():
            continue
        elif skipBecauseFirstLine:
            skipBecauseFirstLine = False
            continue
        try:
            timeStr, value = l.split(";")
            timestamp = dt.datetime.strptime(timeStr, NFF_INPUT_TIMEFORMAT)
            cleanFloat = value.replace(",",".")

            # - means the value is missing in the data set (happens sometimes) #
            if cleanFloat.strip() == "-" or cleanFloat.strip() == "+":
                continue

            dataList += [Data(dtype, float(cleanFloat), timestamp)]

        except ValueError as e:
            print("Warning: {}".format(e))

    return dataList

class Data:
    def __init__(self, dtype, value, time):
        self.dtype = dtype
        self.value = value
        self.time  = time

    def __str__(self):
        return "Data: {} {} {}".format(self.dtype, self.time, self.value)

def sort_func(s):
    title = s.title.replace("März", calendar.month_name[3]) # fix german ä
    return datetime.datetime.strptime(title, "Wetterdaten-%B-%Y.csv")

if __name__ == "__main__":

    # parse arguments #
    parser = argparse.ArgumentParser(description='Reference Data Collector DWD')
    parser.add_argument('--target-file', default="Weatherdata.xlsx", help='File(-path) to save to')
    args = parser.parse_args()

    # check laste months #
    checkLastMonths()

    # read in csv's #
    globPattern = "{}/*.csv".format(CSV_DIR)
    sheets = {}
    for f in glob.glob(globPattern):
        sheet = pyexcel.get_sheet(file_name=f, delimiter=";")
        sheets.update({ os.path.basename(f) : sheet })

    # open & save spreadsheet book #
    book = pyexcel.get_book(bookdict=sheets)
    outfileRaw = args.target_file
    book.save_as(outfileRaw)

    # formating and style #
    wb = openpyxl.load_workbook(filename=outfileRaw)
    longColStart = 5
    for ws in wb.worksheets:

        # width #
        for col in ws.columns:
            width = 20
            longColStart -= 1
            if longColStart < 0:
                width = 40
            ws.column_dimensions[col[0].column_letter].width = 20


        # insert month info row
        ws.insert_rows(1)
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = ws.title[len("Wetterdaten-"):-4].replace("-"," ")
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.fill = openpyxl.styles.PatternFill(start_color='7F03ADFC',
                                        end_color='7F03ADFC', fill_type = 'solid')
        cell.font = openpyxl.styles.Font(bold=True)
        ws.row_dimensions[1].height = 30

        # row height of header (second row behind title) #
        ws.row_dimensions[2].height = 55

        # color / wrap_text / bold #
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='top',
                                                            wrapText=True)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color='7F03ADFC',
                                                end_color='7F03ADFC', fill_type = 'solid')

        # date format #
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=1):
            for cell in row:
                cell.style = DATE_STYLE

        # center everything #
        for row in ws.iter_rows(min_row=3, min_col=1):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    # sort the sheets #
    wb._sheets.sort(key=sort_func, reverse=True)

    wb.save(outfileRaw)
